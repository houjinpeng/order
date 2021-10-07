import html
import re
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import xlwt
import time
#设置表格样式
def set_style(name,height,bold=False):
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = name
    font.bold = bold
    font.color_index = 4
    font.height = height
    style.font = font
    return style

def send_illegal_str(content):
    html = ILLEGAL_CHARACTERS_RE.sub(r'', str(content))

    def str_to_int(s, default, base=10):
        if int(s, base) < 0x10000:
            return chr(int(s, base))
        return default

    html = re.sub(u"&#(\d+);?", lambda c: str_to_int(c.group(1), c.group(0)), html)
    html = re.sub(u"&#[xX]([0-9a-fA-F]+);?", lambda c: str_to_int(c.group(1), c.group(0), base=16), html)
    html = re.sub(u"[\x00-\x08\x0b\x0e-\x1f\x7f]", "", html)
    return html

def write_excel(rows):
    wb = Workbook()
    ws = wb.active
    row0 = ['域名', '网站标题', '使用年龄', '主要语言', '最老记录时间', '最新记录时间', '域名年龄', '标题敏感词检测']
    # 写第一行
    for i in range(0, len(row0)):
        ws.cell(row=1, column=i + 1, value=send_illegal_str(row0[i]))

    row_mount = 2
    for row in rows:
        try:
            ym = row['ym']

            if row['code'] == -1:
                ws.cell(row=row_mount, column=1, value=send_illegal_str(ym))
                row_mount += 1
                continue

            # 写入域名
            ws.cell(row=row_mount, column=1, value=send_illegal_str(ym))
            # 写入网站标题
            ws.cell(row=row_mount, column=2, value=send_illegal_str(row['title']))
            # 检测语言
            # sheet1.write_merge(row_mount, mb_row, 5, 5, row['lang'])
            # 使用年龄
            ws.cell(row=row_mount, column=3, value=send_illegal_str(row['nl']))
            # 主要语言
            ws.cell(row=row_mount, column=4, value=send_illegal_str(row['yy']))
            # 最老记录
            ws.cell(row=row_mount, column=5, value=send_illegal_str(row['sj_min']))
            # 最新记录
            ws.cell(row=row_mount, column=6, value=send_illegal_str(row['sj_max']))
            # 域名年龄
            ws.cell(row=row_mount, column=7, value=send_illegal_str(row['ym_nl']))
            # 标题敏感词
            ws.cell(row=row_mount, column=8, value=send_illegal_str(row['title_sensitive']))
            row_mount += 1
        except Exception as e:
            print("定位到出错行数: ", str(row))
            print(e)
            exit()
    wb.save(f'{int(time.time())}.csv')
    wb.close()


def write_excel_2(rows):
    f = xlwt.Workbook()
    sheet1 = f.add_sheet('域名信息',cell_overwrite_ok=True)
    row0 = ['域名','网站标题','使用年龄','主要语言','最老记录时间','最新记录时间','域名年龄','标题敏感词检测']
    #写第一行
    for i in range(0,len(row0)):
        sheet1.write(0,i,row0[i],set_style('Times New Roman',220,True))

    row_mount = 1

    for row in rows:
        ym = row['ym']

        if row['code'] == -1:
            sheet1.write(row_mount, 0, ym)
            row_mount += 1
            continue
        mb_row = row_mount + len(row['date_list']) - 1
        mb_row = 1 + row_mount
        # 写入域名
        sheet1.write_merge(row_mount, mb_row, 0, 0, ym)
        # 写入网站标题
        sheet1.write_merge(row_mount, mb_row, 1, 1, row['title'])
        #检测语言
        # sheet1.write_merge(row_mount, mb_row, 5, 5, row['lang'])
        #使用年龄
        sheet1.write_merge(row_mount, mb_row, 2, 2, row['nl'])
        #主要语言
        sheet1.write_merge(row_mount, mb_row, 3, 3, row['yy'])
        #最老记录
        sheet1.write_merge(row_mount, mb_row, 4, 4, row['sj_min'])
        #最新记录
        sheet1.write_merge(row_mount, mb_row, 5, 5, row['sj_max'])
        #域名年龄
        sheet1.write_merge(row_mount, mb_row, 6, 6, row['ym_nl'])
        #标题敏感词
        sheet1.write_merge(row_mount, mb_row, 7, 7, row['title_sensitive'])
        #页面敏感词
        # sheet1.write_merge(row_mount, mb_row, 12, 12, row['page_sensitive'])

        # 写入存档时间
        # for date_list in row['date_list']:
        #     sheet1.write(row_mount, 2, date_list['timestamp'])
        #     sheet1.write(row_mount, 3, html.unescape(date_list['bt']))
        #     sheet1.write(row_mount, 4, date_list['yy'])
        row_mount += 1

    f.save(f'{int(time.time())}.csv')
